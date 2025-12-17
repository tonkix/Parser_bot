import logging
import openpyxl
from aiogram import F, Bot, Router
from aiogram.enums import ContentType
from aiogram.types import Message, FSInputFile
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from dotenv import load_dotenv
import os
import time
from datetime import datetime
from openpyxl import Workbook

import app.db.requests as rq
from app.parser_1 import parsing, parsing_one

router = Router()
load_dotenv()
ADMIN_ROLE = os.getenv("ADMIN_ROLE")


class ImportTT(StatesGroup):
    file = State()


def printSheets(sheets):
    list_of_sheets = list()
    for s in sheets:
        list_of_sheets.append(s)
        print(str(list_of_sheets.index(s)) + " | " + s)


def chooseSheet(wb):
    printSheets(wb.sheetnames)
    print('Введите индекс листа: ')
    sheet_index = input()
    ws = wb.sheetnames[int(sheet_index)]
    return ws


def tryDefaultSheetName(wb_data, name):
    try:
        return wb_data[name]
    except Exception as err:
        print(f"Unexpected {err=}, {type(err)=}")
        logging.error(f"sheet name error")
        try:
            return wb_data[str(name).lower()]
        except Exception as err:
            print(f"Unexpected {err=}, {type(err)=}")
            logging.error(f"sheet name error")
            try:
                return wb_data[str(name).upper()]
            except Exception as err:
                print(f"Unexpected {err=}, {type(err)=}")
                logging.error(f"sheet name error")
                sheet = wb_data.sheetnames[int(0)]
                # sheet = chooseSheet(wb_data)
    return wb_data[sheet]


async def find_elem_by_url(url, parsing_result):
    for e in parsing_result:
        if url == e[0]:
            return e


# inputFile - файл для парсинга
# return - Excel файл с результатами парсинга
async def Work_With_File(data: Workbook, creating_dictionary_worksheet, writing_to_db):
    default_sheet_name = "Ссылки"
    data = tryDefaultSheetName(wb_data=data, name=default_sheet_name)
    list_url = list()
    id_url_list = list()
    maxCheckRow = data.max_row + 1  # поиск ссылок в строках
    maxCheckColumn = 30  # поиск ссылок в колонках
    for col in range(0, maxCheckColumn):
        for row in range(0, maxCheckRow):
            cell_url = data.cell(row=row + 1, column=col + 1).value
            product_id = data.cell(row=row + 1, column=1).value
            if "http" in str(cell_url) and "." in str(cell_url):
                list_url.append(cell_url)
                id_url_list.append([product_id, cell_url])

    wb = Workbook()
    ws = wb.active
    ws.title = "Output data"
    uniq_url = list(set(list_url))
    parsing_result = await parsing(uniq_url, ws)

    if creating_dictionary_worksheet:
        # Добавление ссылки в базу и в отдельный лист Excel
        sheet_name = 'dictionary'
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        for k in id_url_list:
            try:
                data = await find_elem_by_url(k[1], parsing_result)
                # print(str(k[0]) + "/" + str(k[1]))
                ws.append({1: k[0],
                           2: k[1],
                           3: data[1],
                           4: data[2]})
                if writing_to_db:
                    await rq.add_link(product_id=k[0],
                                      url=k[1],
                                      name=data[1],
                                      price=data[2])
            except Exception as err:
                mes = f"[ERROR] Unexpected {err=}, {type(err)=}"
                # print(mes)
                logging.error(mes)

    return wb


@router.message(Command("start"))
async def cmd_start(message: Message):
    await rq.set_user(tg_id=message.from_user.id,
                      firstname=message.from_user.first_name,
                      lastname=message.from_user.last_name,
                      subscribed=0,
                      role=1,
                      product_search=1,
                      writing_to_db=1,
                      creating_dictionary_worksheet=1)
    await message.answer(f"Бот умеет"
                         f"\n- Искать товары по id"
                         f"\n- Искать товары по коду товара"
                         f"\n- Искать товары по ссылке"
                         f"\n- Искать товары по слову в названии\n\n"
                         f"\n/start - Начать"
                         f"\n/help - Помощь"
                         f"\n/backup - Выгрузка БД"
                         f"\n/clear_log - Очистить логи"
                         f"\n/subscribe - Подписаться на рассылку"
                         f"\n/unsubscribe - Отписаться от рассылки"
                         f"\n/product_search - Поиск товаров через сообщение"
                         f"\n/writing_to_db - запись ссылок в БД"
                         f"\n/creating_dictionary_worksheet - Переключатель записи в Excel отдельного листа"
                         f"\n/import - Импорт товаров ТТ\n\n"
                         f"Можно отправить файл с ссылками и в ответ бот пришлет файл с результатами парсинга")


@router.message(Command("subscribe"))
async def cmd_subscribe(message: Message):
    await rq.set_user(tg_id=message.from_user.id,
                      firstname=message.from_user.first_name,
                      lastname=message.from_user.last_name,
                      subscribed=1,
                      role=1,
                      product_search=1,
                      writing_to_db=1,
                      creating_dictionary_worksheet=1)
    await rq.subscribe(tg_id=message.from_user.id)
    await message.answer("Вы подписались на рассылку", disable_notification=True)


@router.message(Command("unsubscribe"))
async def cmd_unsubscribe(message: Message):
    await rq.set_user(tg_id=message.from_user.id,
                      firstname=message.from_user.first_name,
                      lastname=message.from_user.last_name,
                      subscribed=0,
                      role=1,
                      product_search=1,
                      writing_to_db=1,
                      creating_dictionary_worksheet=1)
    await rq.unsubscribe(tg_id=message.from_user.id)
    await message.answer("Вы отписались от рассылки", disable_notification=True)


@router.message(Command("clear_log"))
async def cmd_clear_log(message: Message):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        with open("logs.log", 'w') as file:
            file.write('')
        logging.info('Очистка логов')
        await message.answer(f"Логи очищены",
                             disable_notification=True)
    else:
        logging.error('Запрос на очистку логов - не прошла проверка пользователя')
        await message.answer(f"У вас нет доступа для выполнения данной команды")


# Переключение флага записи в БД
@router.message(Command("product_search"))
async def switch_product_search(message: Message):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        await rq.switch_product_search(message.from_user.id)
        user = await rq.get_user_by_tg(message.from_user.id)
        logging.info(f"Enable_Writing_to_DataBase = {user.switch_product_search}")
        await message.answer(f"Enable_Writing_to_DataBase = {user.switch_product_search}",
                             disable_notification=True)
    else:
        logging.error('Запрос изменение флага записи в БД - не прошла проверка пользователя')
        await message.answer(f"У вас нет доступа для выполнения данной команды")


# Переключение флага записи в БД
@router.message(Command("writing_to_db"))
async def switch_writing_to_db(message: Message):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        await rq.switch_writing_to_db(message.from_user.id)
        user = await rq.get_user_by_tg(message.from_user.id)
        logging.info(f"Enable_Writing_to_DataBase = {user.writing_to_db}")
        await message.answer(f"Enable_Writing_to_DataBase = {user.writing_to_db}",
                             disable_notification=True)
    else:
        logging.error('Запрос изменение флага записи в БД - не прошла проверка пользователя')
        await message.answer(f"У вас нет доступа для выполнения данной команды")


# Переключение флага создания в Excel отдельного листа с ID и ссылкой
@router.message(Command("creating_dictionary_worksheet"))
async def switch_creating_dictionary_worksheet(message: Message):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        await rq.switch_creating_dictionary_worksheet(message.from_user.id)
        user = await rq.get_user_by_tg(message.from_user.id)
        logging.info(f"Creating_Dictionary_Worksheet = {user.creating_dictionary_worksheet}")
        await message.answer(f"Creating_Dictionary_Worksheet = {user.creating_dictionary_worksheet}",
                             disable_notification=True)
    else:
        logging.error('Запрос изменение флага записи в доп. лист Excel - не прошла проверка пользователя')
        await message.answer(f"У вас нет доступа для выполнения данной команды")


@router.message(Command("backup"))
async def cmd_backup(message: Message):
    logging.info('Запрос backup_db')
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        logging.info('Запрос backup_db - пользователь принят')

        # не отправляет файл, пишет слишком большой
        document = FSInputFile(path='db.sqlite3', filename='backup_db.sqlite3')
        await message.reply_document(document=document)
    else:
        logging.error('Запрос backup_db - не прошла проверка пользователя')
        await message.answer(f"У вас нет доступа для выполнения данной команды")


@router.message(Command("help"))
async def cmd_help(message: Message):
    await message.answer(f"Бот умеет"
                         f"\n- Искать товары по id"
                         f"\n- Искать товары по коду товара"
                         f"\n- Искать товары по ссылке"
                         f"\n- Искать товары по слову в названии\n\n"
                         f"\n/start - Начать"
                         f"\n/help - Помощь"
                         f"\n/backup - Выгрузка БД"
                         f"\n/clear_log - Очистить логи"
                         f"\n/subscribe - Подписаться на рассылку"
                         f"\n/unsubscribe - Отписаться от рассылки"
                         f"\n/product_search - Поиск товаров через сообщение"
                         f"\n/writing_to_db - запись ссылок в БД"
                         f"\n/creating_dictionary_worksheet - Переключатель записи в Excel отдельного листа"
                         f"\n/import - Импорт товаров ТТ\n\n"
                         f"Можно отправить файл с ссылками и в ответ бот пришлет файл с результатами парсинга")


@router.message(Command("import"))
async def cmd_import(message: Message, state: FSMContext):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.role.__str__() == ADMIN_ROLE:
        await state.set_state(ImportTT.file)
        await message.answer(f"Жду файл импорта")
    else:
        await message.answer(f"У вас нет доступа для выполнения данной команды")


async def update_tt_products(wb: Workbook):
    ws = wb.active
    products_list = list()
    products_dict = dict()
    maxCheckRow = ws.max_row + 1  # поиск ссылок в строках
    for row in range(1, maxCheckRow):
        products_dict['product_tt_id'] = ws.cell(row=row + 1, column=1).value
        products_dict['product_tt_code'] = ws.cell(row=row + 1, column=2).value
        products_dict['name_tt'] = ws.cell(row=row + 1, column=3).value
        products_dict['purchase_price_tt'] = ws.cell(row=row + 1, column=4).value
        products_dict['retail_price_tt'] = ws.cell(row=row + 1, column=5).value
        products_dict['url_tt'] = ws.cell(row=row + 1, column=6).value

        products_list.append([products_dict['product_tt_id'],
                              products_dict['product_tt_code'],
                              products_dict['name_tt'],
                              products_dict['purchase_price_tt'],
                              products_dict['retail_price_tt'],
                              products_dict['url_tt']])

    return products_list


@router.message(ImportTT.file)
async def get_import_file(message: Message, state: FSMContext, bot: Bot):
    await message.answer(f"Начал работу с файлом")
    start = time.perf_counter()
    file_id = message.document.file_id
    input_directory = r'import_data/'
    if not os.path.exists(input_directory):
        os.mkdir(input_directory)
    input_name = "import.xlsx"
    input_file = input_directory + input_name
    await Bot.download(bot, file_id, input_file, 120)
    input_file = openpyxl.load_workbook(os.path.join(input_directory, input_name))
    products = await update_tt_products(input_file)
    for product in products:
        # await rq.update_product(*product)
        # print(product)
        await rq.update_product2(product_tt_id=product[0],
                                 product_tt_code=product[1],
                                 name_tt=product[2],
                                 purchase_price_tt=product[3],
                                 retail_price_tt=product[4],
                                 url_tt=product[5])

    print(f"Добавление файлов заняло {time.perf_counter() - start:0.4f} секунд")
    await message.answer(f"Завершено за {time.perf_counter() - start:0.4f} секунд")
    await state.clear()


async def add_tt_products(wb: Workbook):
    default_sheet_name = "товары"
    ws = tryDefaultSheetName(wb_data=wb, name=default_sheet_name)

    product_list = list()
    maxCheckRow = ws.max_row + 1  # поиск ссылок в строках
    for row in range(0, maxCheckRow):
        product_tt_id = ws.cell(row=row + 1, column=1).value
        product_tt_code = ws.cell(row=row + 1, column=2).value
        name_tt = ws.cell(row=row + 1, column=3).value
        url_tt = ws.cell(row=row + 1, column=4).value
        purchase_price_tt = ws.cell(row=row + 1, column=5).value
        retail_price_tt = ws.cell(row=row + 1, column=6).value

        product_list.append([product_tt_id, product_tt_code, name_tt, url_tt, purchase_price_tt, retail_price_tt])
        await rq.add_tt_product2(product_tt_id=product_tt_id,
                                 product_tt_code=product_tt_code,
                                 name_tt=name_tt,
                                 url_tt=url_tt,
                                 purchase_price_tt=purchase_price_tt,
                                 retail_price_tt=retail_price_tt)


# ловит файл и делает парсинг по ссылкам в файле
@router.message(F.content_type == ContentType.DOCUMENT)
async def get_doc(message: Message, bot: Bot):
    await rq.set_user(tg_id=message.from_user.id,
                      firstname=message.from_user.first_name,
                      lastname=message.from_user.last_name,
                      subscribed=0,
                      role=1,
                      product_search=1,
                      writing_to_db=1,
                      creating_dictionary_worksheet=1)

    file_id = message.document.file_id
    input_directory = r'input_data/'
    if not os.path.exists(input_directory):
        os.mkdir(input_directory)
    input_name = "input.xlsm"
    input_file = input_directory + input_name
    await Bot.download(bot, file_id, input_file, 120)

    input_file = openpyxl.load_workbook(input_file)
    start = time.perf_counter()
    await message.answer('Файл обрабатывается...')
    try:
        user = await rq.get_user_by_tg(message.from_user.id)
        wb = await Work_With_File(input_file,
                                  creating_dictionary_worksheet=user.creating_dictionary_worksheet,
                                  writing_to_db=user.writing_to_db)
    except Exception as err:
        await message.answer(f"Во время парсинга возникла непредвиденная ошибка {err}")
        return
    # print(f"Выполнение заняло {time.perf_counter() - start:0.4f} секунд")
    output_directory = r'output_data/'
    if not os.path.exists(output_directory):
        os.mkdir(output_directory)
    output_name = "output.xlsx"
    output_file = datetime.now().strftime(output_directory + output_name)
    wb.save(output_file)

    await message.reply_document(
        document=FSInputFile(
            path=output_file,
            filename=output_name,
        ),
    )


# Поиск товара
async def find_products(text):
    logging.info("[INFO] Поиск по ID")
    products = await rq.get_product_by_tt_id(text.split(' ')[0])
    product_list = list()
    for p in products:
        product_list.append(p)
    if len(list(product_list)) == 0:
        logging.info("[INFO] Поиск по коду")
        products = await rq.get_product_by_tt_code(text.split(' ')[0])
        for p in products:
            product_list.append(p)
        if len(list(product_list)) == 0:
            logging.info("[INFO] Поиск по ссылке")
            products = await rq.get_products_by_link(text)
            for p in products:
                product_list.append(p)
            if len(list(product_list)) == 0:
                logging.info("[INFO] Поиск по названию")
                products = await rq.get_products_by_name(text)
                for p in products:
                    product_list.append(p)

    return product_list


# @router.message(F.text.contains('товар'))
# ловит любое сообщение, которое не прошло фильтры выше
@router.message()
async def get_links(message: Message):
    user = await rq.get_user_by_tg(message.from_user.id)
    if user.product_search:
        products = await find_products(message.text)
        for product in products:
            logging.info("[INFO] Поиск товара")
            print("[INFO] Поиск товара")
            if product is not None:
                try:
                    data = parsing_one(product.url)
                    await message.answer(text=f"Товар: \nid: {product.product_tt_id}"
                                              f"\nКод товара: {product.product_tt_code}"
                                              f"\nНаименование: {product.name}"
                                              f"\nСсылка: {product.url}"
                                              f"\nЗакуп: {product.purchase_price}"
                                              f"\nРозница: {product.retail_price}"
                                              f"\nДата внесения: {product.update_date}"
                                              f"\n\nТекущее наименование: {data['name']}"
                                              f"\n\nТекущая РЦ: {data['price']}",
                                         disable_notification=True,
                                         disable_web_page_preview=True)
                except Exception as err:
                    mes: str = f"Возникла ошибка {err=}, {type(err)=}"
                    await message.answer(text=mes)
                    print(mes)
                    logging.error(mes)
                links = list(await rq.get_links_by_tt_id(product.product_tt_id))
                if len(links) != 0:
                    await message.answer(text=f"Найдено {len(links)}",
                                         disable_notification=True)
                    for link in links:
                        try:
                            data = parsing_one(link.url)
                            await message.answer(text=f"Ссылка: {link.url}\n\n"
                                                      f"Наименование: {data['name']}\n\n"
                                                      f"Цена: {data['price']}\n",
                                                 disable_notification=True,
                                                 disable_web_page_preview=True)
                        except Exception as err:
                            mes: str = f"Возникла ошибка {err=}, {type(err)=}"
                            await message.answer(text=mes)
                            print(mes)
                            logging.error(mes)
                else:
                    await message.answer(text=f"Ссылок не найдено",
                                         disable_notification=True)

            else:
                await message.answer(text="Товар не найден",
                                     disable_notification=True)
    else:
        await message.answer(text="Поиск товаров отключен",
                             disable_notification=True)
